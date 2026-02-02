#!/usr/bin/env python3
"""Export module cards to a deterministic XLSX summary."""
from __future__ import annotations

import json
from pathlib import Path
from typing import Any

from openpyxl import Workbook

ROOT = Path(__file__).resolve().parents[1]
CARDS_DIR = ROOT / "packages" / "specs" / "docs" / "contracts" / "modules" / "cards"
OUTPUT_PATH = ROOT / "outputs" / "module_cards.xlsx"
REQUIRED_KEYS = {"schema_version", "module_id", "summary", "boundaries", "knobs", "risks"}


def _load_json(path: Path) -> Any:
    return json.loads(path.read_text(encoding="utf-8"))


def _ensure_dict(value: Any, label: str) -> dict[str, Any]:
    if not isinstance(value, dict):
        raise ValueError(f"{label} must be an object")
    return value


def _ensure_list(value: Any, label: str) -> list[Any]:
    if not isinstance(value, list):
        raise ValueError(f"{label} must be an array")
    return value


def _format_default(value: Any) -> str:
    return json.dumps(value, sort_keys=True, separators=(",", ":"))


def _one_line(text: str) -> str:
    return " ".join(text.split())


def _load_cards(cards_dir: Path) -> list[tuple[Path, dict[str, Any]]]:
    if not cards_dir.exists():
        raise ValueError(f"cards directory missing: {cards_dir}")
    card_paths = sorted(cards_dir.glob("*.card.json"))
    if not card_paths:
        raise ValueError(f"no module cards found in {cards_dir}")

    cards: list[tuple[Path, dict[str, Any]]] = []
    for path in card_paths:
        card = _ensure_dict(_load_json(path), str(path))
        missing = sorted(REQUIRED_KEYS - set(card.keys()))
        if missing:
            raise ValueError(f"{path}: missing required keys: {missing}")
        cards.append((path, card))
    return cards


def _module_rows(cards: list[tuple[Path, dict[str, Any]]]) -> list[list[str]]:
    rows: list[list[str]] = []
    for _, card in sorted(cards, key=lambda item: str(item[1].get("module_id", ""))):
        module_id = card.get("module_id")
        summary = card.get("summary")
        if not isinstance(module_id, str) or not module_id:
            raise ValueError("module_id must be a non-empty string")
        if not isinstance(summary, str) or not summary:
            raise ValueError(f"{module_id}: summary must be a non-empty string")
        boundaries = _ensure_dict(card.get("boundaries"), f"{module_id}.boundaries")
        inputs = _ensure_list(boundaries.get("inputs"), f"{module_id}.boundaries.inputs")
        outputs = _ensure_list(boundaries.get("outputs"), f"{module_id}.boundaries.outputs")
        contract_paths: set[str] = set()
        for entry in inputs + outputs:
            entry_dict = _ensure_dict(entry, f"{module_id}.boundaries entry")
            contract_ref = entry_dict.get("contract_ref_path")
            if not isinstance(contract_ref, str) or not contract_ref:
                raise ValueError(f"{module_id}: contract_ref_path must be a non-empty string")
            contract_paths.add(contract_ref)
        contract_list = ", ".join(sorted(contract_paths))
        rows.append([module_id, _one_line(summary), contract_list])
    return rows


def _knob_rows(cards: list[tuple[Path, dict[str, Any]]]) -> list[list[str]]:
    rows: list[list[str]] = []
    for path, card in sorted(cards, key=lambda item: str(item[1].get("module_id", ""))):
        module_id = card.get("module_id")
        if not isinstance(module_id, str) or not module_id:
            raise ValueError("module_id must be a non-empty string")
        knobs = _ensure_list(card.get("knobs"), f"{module_id}.knobs")
        for knob in sorted(knobs, key=lambda item: str(item.get("id", ""))):
            knob_dict = _ensure_dict(knob, f"{module_id}.knob")
            knob_id = knob_dict.get("id")
            value_type = knob_dict.get("value_type")
            required = knob_dict.get("required")
            description = knob_dict.get("description")
            if not isinstance(knob_id, str) or not knob_id:
                raise ValueError(f"{module_id}: knob id must be a non-empty string")
            if not isinstance(value_type, str) or not value_type:
                raise ValueError(f"{module_id}:{knob_id}: value_type must be a string")
            if not isinstance(required, bool):
                raise ValueError(f"{module_id}:{knob_id}: required must be boolean")
            if not isinstance(description, str) or not description:
                raise ValueError(f"{module_id}:{knob_id}: description must be a non-empty string")
            default_value = ""
            if "default" in knob_dict:
                default_value = _format_default(knob_dict["default"])
            rows.append(
                [
                    module_id,
                    knob_id,
                    value_type,
                    "true" if required else "false",
                    default_value,
                    _one_line(description),
                    path.name,
                    "true",
                ]
            )
    return rows


def _io_rows(cards: list[tuple[Path, dict[str, Any]]]) -> list[list[str]]:
    rows: list[list[str]] = []
    for _, card in sorted(cards, key=lambda item: str(item[1].get("module_id", ""))):
        module_id = card.get("module_id")
        if not isinstance(module_id, str) or not module_id:
            raise ValueError("module_id must be a non-empty string")
        boundaries = _ensure_dict(card.get("boundaries"), f"{module_id}.boundaries")
        for direction in ("input", "output"):
            entries = boundaries.get("inputs" if direction == "input" else "outputs")
            for entry in sorted(_ensure_list(entries, f"{module_id}.{direction}s"), key=lambda item: str(item.get("name", ""))):
                entry_dict = _ensure_dict(entry, f"{module_id}.{direction}")
                name = entry_dict.get("name")
                contract_ref = entry_dict.get("contract_ref_path")
                if not isinstance(name, str) or not name:
                    raise ValueError(f"{module_id}: {direction} name must be a non-empty string")
                if not isinstance(contract_ref, str) or not contract_ref:
                    raise ValueError(f"{module_id}: {direction} contract_ref_path must be a non-empty string")
                rows.append([module_id, direction, name, contract_ref])
    return rows


def export_module_cards_to_xlsx(output_path: Path) -> None:
    cards = _load_cards(CARDS_DIR)

    workbook = Workbook()
    default_sheet = workbook.active
    workbook.remove(default_sheet)

    modules_sheet = workbook.create_sheet("Modules")
    modules_sheet.append(["module_id", "summary", "boundary_contract_paths"])
    for row in _module_rows(cards):
        modules_sheet.append(row)

    knobs_sheet = workbook.create_sheet("Knobs")
    knobs_sheet.append(
        [
            "module_id",
            "knob_id",
            "type",
            "required",
            "default",
            "description",
            "source",
            "exposed_in_dashboard",
        ]
    )
    for row in _knob_rows(cards):
        knobs_sheet.append(row)

    io_sheet = workbook.create_sheet("IO")
    io_sheet.append(["module_id", "direction", "name", "contract_ref_path"])
    for row in _io_rows(cards):
        io_sheet.append(row)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)


def main() -> int:
    try:
        export_module_cards_to_xlsx(OUTPUT_PATH)
    except (ValueError, json.JSONDecodeError) as exc:
        print(f"ERROR: {exc}")
        return 1
    print(f"PASS: wrote module cards XLSX to {OUTPUT_PATH}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
